from selenium import webdriver
from selenium.common.exceptions import *
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver import ActionChains
from selenium.webdriver.common.actions.action_builder import ActionBuilder
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support import ui
from time import sleep
from threading import Thread
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
import json

def Find_Element(driver : webdriver.Chrome, by, value : str) -> WebElement:
    while True:
        try:
            element = driver.find_element(by, value)
            break
        except:
            pass
        sleep(0.1)
    return element

def Find_Elements(driver : webdriver.Chrome, by, value : str) -> list[WebElement]:
    while True:
        try:
            elements = driver.find_elements(by, value)
            if len(elements) > 0:
                break
        except:
            pass
        sleep(0.1)
    return elements

def Send_Keys(element : WebElement, content : str):
    element.clear()
    for i in content:
        element.send_keys(i)
        sleep(0.1)

service = Service(executable_path="C:\chromedriver-win64\chromedriver.exe")   
options = Options()
options.add_experimental_option("debuggerAddress", "127.0.0.1:9030")
driver = webdriver.Chrome(service=service, options=options)
driver.get('https://www.worldometers.info/geography/alphabetical-list-of-countries/')

wb = Workbook()
sheet = wb.active

border_style = Side(border_style = "thin", color = "000000")
font = Font(name = 'Times New Roman', size = '12')
alignment = Alignment(vertical = 'center', horizontal = 'center')

item = ["Country", "Population", "Land Area", "Density"]

for i in range(0, 4):
    sheet.cell(row = 1, column = i + 1).value = item[i]
    sheet.cell(row = 1, column = i + 1).border = Border(right = border_style, bottom = border_style)
    sheet.cell(row = 1, column = i + 1).font = font
    sheet.cell(row = 1, column = i + 1).alignment = alignment

countries = Find_Element(driver, By.TAG_NAME, 'tbody').find_elements(By.TAG_NAME, 'tr')
output = []

start_row = 2
for country in countries:
    splits = country.text.split(' ')
    country_name = " ".join(splits[1 : len(splits) - 3])
    output.append({"country" : country_name})
    print(country_name)
    sheet.cell(row = start_row, column = 1).value = country_name
    sheet.cell(row = start_row, column = 1).border = Border(right = border_style, bottom = border_style)
    sheet.cell(row = start_row, column = 1).font = font
    sheet.cell(row = start_row, column = 1).alignment = alignment

    population = country.find_element(By.TAG_NAME, 'a').text
    print(population)
    sheet.cell(row = start_row, column = 2).value = population
    sheet.cell(row = start_row, column = 2).border = Border(right = border_style, bottom = border_style)
    sheet.cell(row = start_row, column = 2).font = font
    sheet.cell(row = start_row, column = 2).alignment = alignment
    
    land_area = splits[len(splits) - 2]
    print(land_area)
    sheet.cell(row = start_row, column = 3).value = land_area
    sheet.cell(row = start_row, column = 3).border = Border(right = border_style, bottom = border_style)
    sheet.cell(row = start_row, column = 3).font = font
    sheet.cell(row = start_row, column = 3).alignment = alignment

    density = splits[len(splits) - 1]
    print(density)
    sheet.cell(row = start_row, column = 4).value = density
    sheet.cell(row = start_row, column = 4).border = Border(right = border_style, bottom = border_style)
    sheet.cell(row = start_row, column = 4).font = font
    sheet.cell(row = start_row, column = 4).alignment = alignment

    start_row += 1

with open('output.json', 'w') as file:
    json.dump(output, file)

wb.save('output.xlsx')